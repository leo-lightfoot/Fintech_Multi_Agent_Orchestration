"""Excel tool -- reads .xlsx and .csv files and returns a JSON summary.

Safety:
- Path is validated against an allowed base directory (DATA_DIR)
- File size capped at 50 MB
- Row count capped at 100,000
- Only .xlsx and .csv extensions accepted

The tool returns a JSON string containing:
  - sheet names (xlsx) or a single "Sheet1"
  - column headers
  - up to MAX_ROWS rows per sheet
  - row and column counts
"""
import json
from pathlib import Path

from langchain_core.tools import tool

from src.utils.config import settings
from src.utils.logging import get_logger

logger = get_logger(__name__)

MAX_FILE_BYTES = 50 * 1024 * 1024   # 50 MB
MAX_ROWS = 100_000


# Project root = two levels up from this file (src/tools/excel.py -> root)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent


def _safe_path(filename: str) -> Path:
    """Resolve filename relative to the project data/ directory.

    Anchored to the project root so it works regardless of CWD.
    Blocks directory traversal outside data/.
    """
    data_base = (_PROJECT_ROOT / "data").resolve()

    candidate = (data_base / filename).resolve()
    try:
        candidate.relative_to(data_base)
    except ValueError:
        raise ValueError(f"Path traversal blocked: {filename}")
    return candidate


@tool
def excel_query(filename: str, sheet: str = "") -> str:
    """Read an Excel (.xlsx) or CSV (.csv) file and return its contents as JSON.

    Args:
        filename: File name relative to the data/ directory (e.g. 'sample_portfolio.xlsx').
        sheet: Sheet name to read (xlsx only). Leave blank to read all sheets.

    Returns:
        JSON string with keys: sheets -> list of {name, columns, rows, total_rows}.
    """
    return _read_file(filename, sheet)


def _read_file(filename: str, sheet: str = "") -> str:
    """Core implementation -- callable directly in tests."""
    try:
        path = _safe_path(filename)
    except ValueError as exc:
        return json.dumps({"error": str(exc)})

    if not path.exists():
        return json.dumps({"error": f"File not found: {filename}"})

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".csv"}:
        return json.dumps({"error": f"Unsupported file type: {suffix}. Use .xlsx or .csv"})

    size = path.stat().st_size
    if size > MAX_FILE_BYTES:
        return json.dumps({"error": f"File too large ({size // (1024*1024)} MB). Limit is 50 MB."})

    try:
        if suffix == ".csv":
            return _read_csv(path)
        return _read_xlsx(path, sheet)
    except Exception as exc:
        logger.error("excel_read_failed", filename=filename, error=str(exc))
        return json.dumps({"error": str(exc)})


def _read_xlsx(path: Path, target_sheet: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames if not target_sheet else [target_sheet]

    sheets_out = []
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = [
            {headers[i]: _safe_val(row[i]) for i in range(len(headers))}
            for row in rows[1: MAX_ROWS + 1]
        ]
        sheets_out.append({
            "name": name,
            "columns": headers,
            "rows": data_rows,
            "total_rows": ws.max_row - 1,
        })
    wb.close()
    logger.info("excel_read_ok", file=path.name, sheets=len(sheets_out))
    return json.dumps({"sheets": sheets_out}, default=str)


def _read_csv(path: Path) -> str:
    import csv
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            if i >= MAX_ROWS:
                break
            rows.append(dict(row))
    result = {"sheets": [{"name": "Sheet1", "columns": list(headers), "rows": rows, "total_rows": len(rows)}]}
    logger.info("csv_read_ok", file=path.name, rows=len(rows))
    return json.dumps(result, default=str)


def _safe_val(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    return str(v)
